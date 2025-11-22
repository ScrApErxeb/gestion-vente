"""
Routes API supplémentaires
"""
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required, current_user
from models import Notification, User, db, Vente
from utils.export import exporter_ventes_pdf, exporter_produits_excel
from utils.helpers import convertir_devise, get_system_parameter, set_system_parameter
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

api_bp = Blueprint('api', __name__)


@api_bp.route('/dashboard')
@login_required
def api_dashboard():
    """Endpoint pour les données du dashboard"""
    try:
        # Données simulées - À REMPLACER par vos vraies données
        stats = {
            'devise': 'F CFA',
            'ventes': {
                'jour': 125000,
                'mois': 1850000,
                'annee': 22450000,
                'mensuelles': [
                    {'mois': 'Jan', 'montant': 1200000},
                    {'mois': 'Fév', 'montant': 1500000},
                    {'mois': 'Mar', 'montant': 1800000},
                    {'mois': 'Avr', 'montant': 1650000},
                    {'mois': 'Mai', 'montant': 1900000},
                    {'mois': 'Jun', 'montant': 2100000},
                    {'mois': 'Jul', 'montant': 1950000},
                    {'mois': 'Aoû', 'montant': 1850000},
                    {'mois': 'Sep', 'montant': 2000000},
                    {'mois': 'Oct', 'montant': 2200000},
                    {'mois': 'Nov', 'montant': 2350000},
                    {'mois': 'Déc', 'montant': 1850000}
                ]
            },
            'stats_globales': {
                'categories': 8,
                'fournisseurs': 12
            },
            'produits': {
                'total': 156,
                'stock_faible': 7,
                'top_ventes': [
                    {'nom': 'Smartphone X10', 'quantite': 45, 'ca': 2250000},
                    {'nom': 'Ecouteurs Pro', 'quantite': 38, 'ca': 950000},
                    {'nom': 'Chargeur Rapide', 'quantite': 32, 'ca': 320000},
                    {'nom': 'Tablette Mini', 'quantite': 28, 'ca': 1960000},
                    {'nom': 'Powerbank 20000mAh', 'quantite': 25, 'ca': 625000},
                    {'nom': 'Câble USB-C', 'quantite': 22, 'ca': 110000},
                    {'nom': 'Haut-parleur Bluetooth', 'quantite': 18, 'ca': 720000}
                ]
            },
            'clients': {
                'total': 124,
                'nouveaux': 8
            },
            'commandes': {
                'en_cours': 5,
                'montant_en_cours': 750000
            }
        }
        
        return jsonify(stats)
        
    except Exception as e:
        print(f"Erreur API dashboard: {e}")
        return jsonify({'error': str(e)}), 500

@api_bp.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications():
    notifications = Notification.query.filter_by(
        user_id=current_user.id
    ).order_by(Notification.date_creation.desc()).limit(50).all()
    
    return jsonify([n.to_dict() for n in notifications])

@api_bp.route('/api/notifications/<int:id>/lire', methods=['POST'])
@login_required
def marquer_notification_lue(id):
    notif = Notification.query.get_or_404(id)
    notif.lue = True
    db.session.commit()
    return jsonify({'success': True})

@api_bp.route('/api/notifications/marquer-toutes-lues', methods=['POST'])
@login_required
def marquer_toutes_notifications_lues():
    Notification.query.filter_by(
        user_id=current_user.id,
        lue=False
    ).update({'lue': True})
    db.session.commit()
    return jsonify({'success': True})

@api_bp.route('/api/devise/convertir', methods=['POST'])
@login_required
def convertir_devise_api():
    """Convertit un montant d'une devise à une autre"""
    data = request.json
    montant = data['montant']
    devise_source = data['devise_source']
    devise_cible = data['devise_cible']
    
    from utils.helpers import convertir_devise
    montant_converti = convertir_devise(montant, devise_source, devise_cible)
    
    return jsonify({
        'montant': montant_converti,
        'montant_source': montant,
        'devise_source': devise_source,
        'montant_cible': montant_converti,
        'devise_cible': devise_cible
    })

@api_bp.route('/api/export/ventes/pdf')
@login_required
def export_ventes_pdf():
    """Exporte les ventes en PDF"""
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    
    query = Vente.query.filter_by(statut='confirmée')
    
    if date_debut:
        query = query.filter(Vente.date_vente >= datetime.fromisoformat(date_debut))
    if date_fin:
        query = query.filter(Vente.date_vente <= datetime.fromisoformat(date_fin))
    
    ventes = query.order_by(Vente.date_vente.desc()).all()
    pdf_buffer = exporter_ventes_pdf(ventes)
    
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f'ventes_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
        mimetype='application/pdf'
    )

@api_bp.route('/api/export/produits/excel')
@login_required
def export_produits_excel():
    """Exporte les produits en Excel"""
    from models import Produit
    
    produits = Produit.query.filter_by(actif=True).all()
    excel_buffer = exporter_produits_excel(produits)
    
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=f'produits_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@api_bp.route('/api/users', methods=['GET'])
@login_required
def liste_utilisateurs_api():
    """Retourne la liste des utilisateurs"""
    if current_user.role != 'admin':
        return jsonify({'message': 'Accès non autorisé'}), 403
        
    users = User.query.all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'email': u.email,
        'nom': u.nom,
        'prenom': u.prenom,
        'role': u.role,
        'telephone': u.telephone,
        'actif': u.actif,
        'date_creation': u.date_creation.isoformat() if u.date_creation else None,
        'dernier_login': u.dernier_login.isoformat() if u.dernier_login else None
    } for u in users])

@api_bp.route('/api/system/parametres', methods=['GET'])
@login_required
def get_parametres_systeme():
    """Récupère les paramètres système"""
    if current_user.role != 'admin':
        return jsonify({'message': 'Accès non autorisé'}), 403
        
    from models.parametre_systeme import ParametreSysteme
    parametres = ParametreSysteme.query.all()
    
    return jsonify([{
        'cle': p.cle,
        'valeur': p.valeur,
        'description': p.description,
        'type_valeur': p.type_valeur
    } for p in parametres])

@api_bp.route('/api/system/parametres', methods=['POST'])
@login_required
def set_parametres_systeme():
    """Définit les paramètres système"""
    if current_user.role != 'admin':
        return jsonify({'message': 'Accès non autorisé'}), 403
        
    data = request.json
    
    for param in data:
        set_system_parameter(
            param['cle'],
            param['valeur'],
            param.get('type_valeur', 'string')
        )
    
    return jsonify({'success': True})

@api_bp.route('/api/users', methods=['POST'])
@login_required
def ajouter_utilisateur_api():
    """Ajoute un nouvel utilisateur"""
    try:
        # Vérifier les permissions
        if current_user.role != 'admin':
            return jsonify({'message': 'Accès non autorisé'}), 403
            
        data = request.json
        print(f"📝 Données création utilisateur: {data}")
        
        # Vérification des données requises
        required_fields = ['username', 'email', 'password', 'nom', 'prenom', 'role']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'message': f'Le champ {field} est requis'}), 400
        
        # Vérification si username/email existe déjà
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'message': 'Ce nom d\'utilisateur existe déjà'}), 400
        
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'message': 'Cet email existe déjà'}), 400
        
        # Création du nouvel utilisateur
        new_user = User(
            username=data['username'],
            email=data['email'],
            nom=data['nom'],
            prenom=data['prenom'],
            role=data['role'],
            telephone=data.get('telephone'),
            actif=data.get('actif', True)
        )
        new_user.set_password(data['password'])
        
        db.session.add(new_user)
        db.session.commit()
        
        print(f"✅ Utilisateur créé: {new_user.username}")
        return jsonify({
            'message': 'Utilisateur créé avec succès',
            'user': new_user.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur création utilisateur: {e}")
        return jsonify({'message': 'Erreur lors de la création: ' + str(e)}), 500

@api_bp.route('/api/users/<int:user_id>/toggle-status', methods=['POST'])
@login_required
def toggle_user_status(user_id):
    try:
        user = User.query.get_or_404(user_id)
        user.actif = not user.actif
        db.session.commit()
        
        return jsonify({
            'message': f'Utilisateur {"activé" if user.actif else "désactivé"} avec succès',
            'user_id': user.id,
            'is_active': user.actif,
            'username': user.username
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api_bp.route('/api/export/ventes/excel', methods=['GET'])
@login_required
def export_ventes_excel_route():
    """Exporte les ventes en Excel (version sans pandas)"""
    try:
        # Récupérer les paramètres de date
        date_debut = request.args.get('date_debut')
        date_fin = request.args.get('date_fin')
        
        # Valider les paramètres
        if not date_debut or not date_fin:
            return jsonify({'error': 'Les paramètres date_debut et date_fin sont requis'}), 400
        
        # Convertir les dates
        try:
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d')
            date_fin_obj = date_fin_obj + timedelta(days=1)
        except ValueError:
            return jsonify({'error': 'Format de date invalide. Utilisez YYYY-MM-DD'}), 400
        
        print(f"🔍 Recherche ventes du {date_debut} au {date_fin}")
        
        # Récupérer les ventes dans la période
        ventes = Vente.query.filter(
            Vente.date_vente >= date_debut_obj,
            Vente.date_vente < date_fin_obj
        ).all()
        
        print(f"📊 {len(ventes)} ventes trouvées")
        
        # Créer le fichier Excel avec openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventes"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-têtes
        headers = [
            'ID', 'Numéro Facture', 'Date', 'Client', 'Produit', 
            'Quantité', 'Prix Unitaire', 'Remise', 'Montant Total', 
            'Devise', 'Mode Paiement', 'Statut', 'Statut Paiement'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données des ventes
        for row, vente in enumerate(ventes, 2):
            vente_data = vente.to_dict()
            ws.cell(row=row, column=1, value=vente.id)
            ws.cell(row=row, column=2, value=vente.numero_facture)
            ws.cell(row=row, column=3, value=vente.date_vente.strftime('%Y-%m-%d %H:%M') if vente.date_vente else '')
            ws.cell(row=row, column=4, value=vente_data.get('client', 'Client anonyme'))
            ws.cell(row=row, column=5, value=vente_data.get('produit', 'Produit inconnu'))
            ws.cell(row=row, column=6, value=vente.quantite)
            ws.cell(row=row, column=7, value=float(vente.prix_unitaire))
            ws.cell(row=row, column=8, value=float(vente.remise))
            ws.cell(row=row, column=9, value=float(vente.montant_total))
            ws.cell(row=row, column=10, value=vente.devise)
            ws.cell(row=row, column=11, value=vente.mode_paiement)
            ws.cell(row=row, column=12, value=vente.statut)
            ws.cell(row=row, column=13, value=vente.statut_paiement)
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder dans le buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"ventes_{date_debut}_a_{date_fin}.xlsx"
        print(f"✅ Export Excel réussi: {filename}")
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Erreur export Excel: {str(e)}")
        return jsonify({'error': f'Erreur lors de l\'export: {str(e)}'}), 500   

@api_bp.route('/api/export/ventes/preview', methods=['GET'])
@login_required
def preview_ventes_export():
    """Aperçu des données qui seront exportées"""
    try:
        date_debut = request.args.get('date_debut')
        date_fin = request.args.get('date_fin')
        
        if not date_debut or not date_fin:
            return jsonify({'error': 'Les paramètres date_debut et date_fin sont requis'}), 400
        
        # Même logique de récupération que l'export
        from models.produit import Produit
        from models.client import Client
        
        ventes = db.session.query(
            Vente, Produit, Client
        ).join(
            Produit, Vente.produit_id == Produit.id
        ).outerjoin(
            Client, Vente.client_id == Client.id
        ).filter(
            Vente.date_vente >= datetime.strptime(date_debut, '%Y-%m-%d'),
            Vente.date_vente < datetime.strptime(date_fin, '%Y-%m-%d') + timedelta(days=1)
        ).all()
        
        preview_data = []
        for vente, produit, client in ventes[:5]:  # Limiter à 5 pour l'aperçu
            preview_data.append({
                'numero_facture': vente.numero_facture,
                'date': vente.date_vente.strftime('%Y-%m-%d %H:%M'),
                'client': f"{client.nom} {client.prenom}" if client else "Client anonyme",
                'produit': produit.nom,
                'quantite': vente.quantite,
                'montant_total': vente.montant_total
            })
        
        return jsonify({
            'period': f"{date_debut} à {date_fin}",
            'total_ventes': len(ventes),
            'preview': preview_data,
            'export_url': f"/api/export/ventes/excel?date_debut={date_debut}&date_fin={date_fin}"
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    

@api_bp.route('/api/export/all-data', methods=['GET'])
@login_required
def export_all_data():
    """Exporte toutes les données de l'application en Excel (sans pandas)"""
    try:
        # Vérifier les permissions admin
        if current_user.role != 'admin':
            return jsonify({'message': 'Accès non autorisé'}), 403
        
        print("🔄 Début de l'export de toutes les données...")
        
        # Créer un fichier Excel avec openpyxl
        wb = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut si elle existe
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # === FEUILLE 1: UTILISATEURS ===
        users = User.query.all()
        if users:
            ws_users = wb.create_sheet("Utilisateurs")
            headers_users = [
                'ID', 'Username', 'Email', 'Nom', 'Prénom', 'Rôle', 
                'Téléphone', 'Actif', 'Date Création', 'Dernier Login'
            ]
            
            # En-têtes
            for col, header in enumerate(headers_users, 1):
                cell = ws_users.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Données
            for row, user in enumerate(users, 2):
                ws_users.cell(row=row, column=1, value=user.id)
                ws_users.cell(row=row, column=2, value=user.username)
                ws_users.cell(row=row, column=3, value=user.email)
                ws_users.cell(row=row, column=4, value=user.nom or '')
                ws_users.cell(row=row, column=5, value=user.prenom or '')
                ws_users.cell(row=row, column=6, value=user.role)
                ws_users.cell(row=row, column=7, value=user.telephone or '')
                ws_users.cell(row=row, column=8, value='Oui' if user.actif else 'Non')
                ws_users.cell(row=row, column=9, value=user.date_creation.strftime('%Y-%m-%d %H:%M') if user.date_creation else '')
                ws_users.cell(row=row, column=10, value=user.dernier_login.strftime('%Y-%m-%d %H:%M') if user.dernier_login else '')
            
            print(f"✅ {len(users)} utilisateurs exportés")
        
        # === FEUILLE 2: VENTES ===
        ventes = Vente.query.all()
        if ventes:
            ws_ventes = wb.create_sheet("Ventes")
            headers_ventes = [
                'ID', 'Numéro Facture', 'Date', 'Client', 'Produit', 
                'Quantité', 'Prix Unitaire', 'Remise', 'Montant Total', 
                'Devise', 'Mode Paiement', 'Statut', 'Statut Paiement'
            ]
            
            # En-têtes
            for col, header in enumerate(headers_ventes, 1):
                cell = ws_ventes.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Données
            for row, vente in enumerate(ventes, 2):
                vente_data = vente.to_dict()
                ws_ventes.cell(row=row, column=1, value=vente.id)
                ws_ventes.cell(row=row, column=2, value=vente.numero_facture)
                ws_ventes.cell(row=row, column=3, value=vente.date_vente.strftime('%Y-%m-%d %H:%M') if vente.date_vente else '')
                ws_ventes.cell(row=row, column=4, value=vente_data.get('client', 'Client anonyme'))
                ws_ventes.cell(row=row, column=5, value=vente_data.get('produit', 'Produit inconnu'))
                ws_ventes.cell(row=row, column=6, value=vente.quantite)
                ws_ventes.cell(row=row, column=7, value=float(vente.prix_unitaire))
                ws_ventes.cell(row=row, column=8, value=float(vente.remise))
                ws_ventes.cell(row=row, column=9, value=float(vente.montant_total))
                ws_ventes.cell(row=row, column=10, value=vente.devise)
                ws_ventes.cell(row=row, column=11, value=vente.mode_paiement)
                ws_ventes.cell(row=row, column=12, value=vente.statut)
                ws_ventes.cell(row=row, column=13, value=vente.statut_paiement)
            
            print(f"✅ {len(ventes)} ventes exportées")
        
        # === FEUILLE 3: PRODUITS ===
        try:
            from models.produit import Produit
            produits = Produit.query.all()
            if produits:
                ws_produits = wb.create_sheet("Produits")
                headers_produits = [
                    'ID', 'Nom', 'Référence', 'Description', 'Prix Achat', 
                    'Prix Vente', 'Stock Actuel', 'Stock Min', 'Catégorie', 
                    'Fournisseur', 'Actif'
                ]
                
                # En-têtes
                for col, header in enumerate(headers_produits, 1):
                    cell = ws_produits.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Données
                for row, produit in enumerate(produits, 2):
                    produit_data = produit.to_dict()
                    ws_produits.cell(row=row, column=1, value=produit.id)
                    ws_produits.cell(row=row, column=2, value=produit.nom)
                    ws_produits.cell(row=row, column=3, value=produit.reference)
                    ws_produits.cell(row=row, column=4, value=produit.description or '')
                    ws_produits.cell(row=row, column=5, value=float(produit.prix_achat))
                    ws_produits.cell(row=row, column=6, value=float(produit.prix_vente))
                    ws_produits.cell(row=row, column=7, value=produit.stock_actuel)
                    ws_produits.cell(row=row, column=8, value=produit.stock_min)
                    ws_produits.cell(row=row, column=9, value=produit_data.get('categorie', ''))
                    ws_produits.cell(row=row, column=10, value=produit_data.get('fournisseur', ''))
                    ws_produits.cell(row=row, column=11, value='Oui' if produit.actif else 'Non')
                
                print(f"✅ {len(produits)} produits exportés")
        except Exception as e:
            print(f"⚠️ Erreur export produits: {e}")
        
        # === FEUILLE 4: STATISTIQUES ===
        ws_stats = wb.create_sheet("Statistiques")
        
        stats_data = [
            ['Statistique', 'Valeur'],
            ['Date de génération', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Utilisateurs', len(users)],
            ['Total Ventes', len(ventes)],
            ['Total Produits', len(produits) if 'produits' in locals() else 'N/A'],
            ['Chiffre d\'affaires total', f"{sum([v.montant_total for v in ventes]):.2f} XOF" if ventes else '0.00 XOF']
        ]
        
        for row, (stat, valeur) in enumerate(stats_data, 1):
            ws_stats.cell(row=row, column=1, value=stat)
            ws_stats.cell(row=row, column=2, value=valeur)
            if row == 1:  # En-tête
                ws_stats.cell(row=row, column=1).font = header_font
                ws_stats.cell(row=row, column=1).fill = header_fill
                ws_stats.cell(row=row, column=2).font = header_font
                ws_stats.cell(row=row, column=2).fill = header_fill
        
        # Ajuster toutes les colonnes
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nom du fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_complet_{timestamp}.xlsx"
        
        print(f"✅ Export complet réussi: {filename}")
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Erreur export complet: {str(e)}")
        return jsonify({'error': f'Erreur lors de l\'export complet: {str(e)}'}), 500    

@api_bp.route('/api/notifications/nettoyer', methods=['POST'])
@login_required
def nettoyer_notifications():
    """Supprime toutes les notifications lues"""
    try:
        if current_user.role != 'admin':
            return jsonify({'message': 'Accès non autorisé'}), 403
        
        # Compter le nombre de notifications avant suppression
        count_avant = Notification.query.filter_by(user_id=current_user.id).count()
        
        # Supprimer les notifications lues
        Notification.query.filter_by(
            user_id=current_user.id,
            lue=True
        ).delete()
        
        db.session.commit()
        
        # Compter après suppression
        count_apres = Notification.query.filter_by(user_id=current_user.id).count()
        notifications_supprimees = count_avant - count_apres
        
        print(f"🗑️ {notifications_supprimees} notifications nettoyées")
        
        return jsonify({
            'success': True,
            'message': f'{notifications_supprimees} notifications supprimées',
            'notifications_supprimees': notifications_supprimees
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur nettoyage notifications: {e}")
        return jsonify({'error': str(e)}), 500
    


# routes/api_system.py
from models import ParametreSysteme  # importe ton modèle


@api_bp.route('/system/control', methods=['POST'])
@login_required
def system_control():
    """Contrôle le système (démarrage/arrêt) avec persistance dans la DB"""
    if current_user.role != 'admin':
        return jsonify({'message': 'Accès non autorisé'}), 403

    data = request.json
    action = data.get('action')

    # État courant du système
    current_status = ParametreSysteme.get_value('system_status', default='stopped')

    if action == 'start':
        if current_status == 'running':
            return jsonify({'success': False, 'message': 'Le système est déjà en cours', 'status': 'running'})
        ParametreSysteme.set_value('system_status', 'running')
        return jsonify({'success': True, 'message': 'Système démarré avec succès', 'status': 'running'})

    elif action == 'stop':
        if current_status == 'stopped':
            return jsonify({'success': False, 'message': 'Le système est déjà arrêté', 'status': 'stopped'})
        ParametreSysteme.set_value('system_status', 'stopped')
        return jsonify({'success': True, 'message': 'Système arrêté avec succès', 'status': 'stopped'})

    elif action == 'restart':
        ParametreSysteme.set_value('system_status', 'running')
        return jsonify({'success': True, 'message': 'Système redémarré avec succès', 'status': 'running'})

    else:
        return jsonify({'error': 'Action non reconnue'}), 400


@api_bp.route('/system/status', methods=['GET'])
@login_required
def system_status():
    """Renvoie l'état actuel du système"""
    if current_user.role != 'admin':
        return jsonify({'message': 'Accès non autorisé'}), 403

    status = ParametreSysteme.get_value('system_status', default='stopped')
    return jsonify({'status': status})

from utils.export import exporter_facture_pdf

@api_bp.route('/export/facture/<int:vente_id>', methods=['GET'])
@login_required
def export_facture_pdf_route(vente_id):
    """Export PDF d'une facture unique"""
    vente = Vente.query.get(vente_id)
    if not vente:
        return jsonify({'error': 'Facture non trouvée'}), 404

    pdf_buffer = exporter_facture_pdf(vente)
    filename = f"facture_{vente.numero_facture}.pdf"

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


