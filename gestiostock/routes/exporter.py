# routes/exporter.py
from flask import Blueprint, request, send_file, jsonify
from flask_login import login_required
from models import Vente
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

exporter_bp = Blueprint('exporter', __name__, url_prefix='/routes/exporter')


@exporter_bp.route('/ventes/excel', methods=['GET'])
@login_required
def export_ventes_excel():
    """Exporte les ventes en Excel"""
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')
    
    if not date_debut or not date_fin:
        return jsonify({'error': 'Les paramètres date_debut et date_fin sont requis'}), 400
    
    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d') + timedelta(days=1)
    except ValueError:
        return jsonify({'error': 'Format de date invalide. Utilisez YYYY-MM-DD'}), 400
    
    ventes = Vente.query.filter(
        Vente.date_vente >= date_debut_obj,
        Vente.date_vente < date_fin_obj
    ).all()
    
    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventes"
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ['ID', 'Numéro Facture', 'Date', 'Client', 'Produit', 'Quantité',
               'Prix Unitaire', 'Remise', 'Montant Total', 'Devise', 'Mode Paiement',
               'Statut', 'Statut Paiement']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row, vente in enumerate(ventes, 2):
        v = vente.to_dict()
        ws.cell(row=row, column=1, value=vente.id)
        ws.cell(row=row, column=2, value=vente.numero_facture)
        ws.cell(row=row, column=3, value=vente.date_vente.strftime('%Y-%m-%d %H:%M') if vente.date_vente else '')
        ws.cell(row=row, column=4, value=v.get('client', 'Client anonyme'))
        ws.cell(row=row, column=5, value=v.get('produit', 'Produit inconnu'))
        ws.cell(row=row, column=6, value=vente.quantite)
        ws.cell(row=row, column=7, value=float(vente.prix_unitaire))
        ws.cell(row=row, column=8, value=float(vente.remise))
        ws.cell(row=row, column=9, value=float(vente.montant_total))
        ws.cell(row=row, column=10, value=vente.devise)
        ws.cell(row=row, column=11, value=vente.mode_paiement)
        ws.cell(row=row, column=12, value=vente.statut)
        ws.cell(row=row, column=13, value=vente.statut_paiement)
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"ventes_{date_debut}_a_{date_fin}.xlsx"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
