from flask import Blueprint, request, jsonify, render_template
from flask_login import login_required, current_user
from sqlalchemy import func
from models import db, User, Fournisseur  # Ajouter cette ligne
from datetime import datetime, timedelta
from models import Vente, Produit, Client, Categorie, VenteItem, Commande, Depense
import json

statistiques_bp = Blueprint('statistiques', __name__)

@statistiques_bp.route('/statistiques')
@login_required
def statistiques_page():
    return render_template('statistiques.html')

@statistiques_bp.route('/api/dashboard')
@login_required
def dashboard_data():
    """Données du tableau de bord avec analyse avancée"""
    today = datetime.now().date()
    start_of_day = datetime.combine(today, datetime.min.time())
    start_of_month = datetime(today.year, today.month, 1)
    start_of_year = datetime(today.year, 1, 1)
    
    # Devise de l'utilisateur
    user_prefs = json.loads(current_user.preferences) if current_user.preferences else {}
    devise = user_prefs.get('devise', 'XOF')
    
    # Ventes
    ventes_jour = db.session.query(func.sum(Vente.montant_total)).filter(
        Vente.date_vente >= start_of_day,
        Vente.statut == 'confirmée'
    ).scalar() or 0
    
    ventes_mois = db.session.query(func.sum(Vente.montant_total)).filter(
        Vente.date_vente >= start_of_month,
        Vente.statut == 'confirmée'
    ).scalar() or 0
    
    ventes_annee = db.session.query(func.sum(Vente.montant_total)).filter(
        Vente.date_vente >= start_of_year,
        Vente.statut == 'confirmée'
    ).scalar() or 0
    
    # Commandes fournisseurs
    commandes_en_cours = Commande.query.filter(
        Commande.statut.in_(['en_attente', 'confirmée'])
    ).count()
    
    montant_commandes_en_cours = db.session.query(func.sum(Commande.montant_total)).filter(
        Commande.statut.in_(['en_attente', 'confirmée'])
    ).scalar() or 0
    
    # Statistiques produits
    nb_produits = Produit.query.filter_by(actif=True).count()
    produits_stock_faible = Produit.query.filter(
        Produit.stock_actuel <= Produit.stock_min,
        Produit.actif == True
    ).count()
    
    valeur_stock_total = db.session.query(
        func.sum(Produit.stock_actuel * Produit.prix_achat)
    ).filter_by(actif=True).scalar() or 0
    
    # Top produits vendus (mois en cours)
    top_produits = (
    db.session.query(
        Produit.nom,
        func.sum(VenteItem.quantite).label('total_vendu'),
        func.sum(VenteItem.montant_total).label('ca')
    )
    .join(VenteItem, Produit.id == VenteItem.produit_id)
    .join(Vente, Vente.id == VenteItem.vente_id)
    .filter(
        Vente.date_vente >= start_of_month,
        Vente.statut == 'confirmée'
    )
    .group_by(Produit.id)
    .order_by(func.sum(VenteItem.quantite).desc())
    .limit(5)
    .all()
)
    # Ventes par mois (12 derniers mois)
    ventes_mensuelles = []
    for i in range(11, -1, -1):
        date = datetime.now() - timedelta(days=30*i)
        debut_mois = datetime(date.year, date.month, 1)
        if date.month == 12:
            fin_mois = datetime(date.year + 1, 1, 1)
        else:
            fin_mois = datetime(date.year, date.month + 1, 1)
        
        total = db.session.query(func.sum(Vente.montant_total)).filter(
            Vente.date_vente >= debut_mois,
            Vente.date_vente < fin_mois,
            Vente.statut == 'confirmée'
        ).scalar() or 0
        
        ventes_mensuelles.append({
            'mois': debut_mois.strftime('%b %Y'),
            'montant': total
        })
    
    # Statistiques clients
    nb_clients = Client.query.filter_by(actif=True).count()
    nouveaux_clients = Client.query.filter(Client.date_creation >= start_of_month).count()
    
    # Top clients
    top_clients = db.session.query(
        Client.nom,
        Client.prenom,
        func.sum(Vente.montant_total).label('total')
    ).join(Vente).filter(
        Vente.date_vente >= start_of_month,
        Vente.statut == 'confirmée'
    ).group_by(Client.id).order_by(func.sum(Vente.montant_total).desc()).limit(5).all()
    
    return jsonify({
        'ventes': {
            'jour': ventes_jour,
            'mois': ventes_mois,
            'annee': ventes_annee,
            'mensuelles': ventes_mensuelles
        },
        'commandes': {
            'en_cours': commandes_en_cours,
            'montant_en_cours': montant_commandes_en_cours
        },
        'produits': {
            'total': nb_produits,
            'stock_faible': produits_stock_faible,
            'valeur_stock': valeur_stock_total,
            'top_ventes': [{'nom': p[0], 'quantite': p[1], 'ca': p[2]} for p in top_produits]
        },
        'clients': {
            'total': nb_clients,
            'nouveaux': nouveaux_clients,
            'top_clients': [{'nom': f"{c[0]} {c[1]}", 'total': c[2]} for c in top_clients]
        },
        'devise': devise,
        'stats_globales': {
            'categories': Categorie.query.filter_by(actif=True).count(),
            'fournisseurs': Fournisseur.query.filter_by(actif=True).count(),
            'utilisateurs': User.query.filter_by(actif=True).count()
        }
    })

# ... (les autres routes statistiques que nous avions corrigées précédemment)
@statistiques_bp.route('/api/stats/ventes')
@login_required
def stats_ventes_complete():
    """Statistiques complètes des ventes avec toutes les données nécessaires"""
    try:
        periode = request.args.get('periode', 'mois')
        now = datetime.now()

        # Calcul de la période
        if periode == 'jour':
            debut = datetime(now.year, now.month, now.day)
        elif periode == 'semaine':
            debut = now - timedelta(days=now.weekday())
        elif periode == 'mois':
            debut = datetime(now.year, now.month, 1)
        elif periode == 'annee':
            debut = datetime(now.year, 1, 1)
        else:
            debut = datetime(now.year, now.month, 1)

        # Ventes totales
        ventes_totales = db.session.query(
            func.sum(Vente.montant_total).label('total'),
            func.count(Vente.id).label('nb')
        ).filter(
            Vente.date_vente >= debut,
            Vente.statut == 'confirmée'
        ).first()

        ca_total = ventes_totales.total or 0
        nb_ventes = ventes_totales.nb or 0
        panier_moyen = ca_total / nb_ventes if nb_ventes else 0

        # Évolution jour par jour (30 derniers jours)
        evolution = []
        for i in range(29, -1, -1):
            date_i = datetime.now() - timedelta(days=i)
            date_debut_jour = datetime(date_i.year, date_i.month, date_i.day)
            date_fin_jour = date_debut_jour + timedelta(days=1)
            
            total_jour = db.session.query(func.sum(Vente.montant_total)).filter(
                Vente.date_vente >= date_debut_jour,
                Vente.date_vente < date_fin_jour,
                Vente.statut == 'confirmée'
            ).scalar() or 0
            
            evolution.append({
                "date": date_i.strftime("%d %b"),
                "montant": total_jour
            })

        # Ventes par catégorie - CORRIGÉ
        ventes_par_categorie = db.session.query(
            Categorie.nom.label('categorie'),
            func.sum(VenteItem.prix_unitaire * VenteItem.quantite * (1 - (VenteItem.remise/100))).label('total')
        ).join(Produit, Produit.categorie_id == Categorie.id)\
        .join(VenteItem, VenteItem.produit_id == Produit.id)\
        .join(Vente, Vente.id == VenteItem.vente_id)\
        .filter(
            Vente.statut == 'confirmée',
            Vente.date_vente >= debut
        ).group_by(Categorie.nom).all()



        # Ventes par mode de paiement - CORRIGÉ
        ventes_par_paiement = db.session.query(
            Vente.mode_paiement,
            func.sum(Vente.montant_total).label('total')
        ).filter(
            Vente.statut == 'confirmée',
            Vente.date_vente >= debut
        ).group_by(Vente.mode_paiement).all()

        # Top 10 produits vendus (quantité & chiffre d'affaires)
        top_produits = db.session.query(
            Produit.nom.label('produit'),
            func.sum(VenteItem.quantite).label('quantite_totale'),
            func.sum(VenteItem.quantite * VenteItem.prix_unitaire * (1 - (VenteItem.remise/100))).label('ca_total')
        ).join(VenteItem, Produit.id == VenteItem.produit_id)\
        .join(Vente, Vente.id == VenteItem.vente_id)\
        .filter(
            Vente.statut == 'confirmée',
            Vente.date_vente >= debut
        ).group_by(Produit.id)\
        .order_by(func.sum(VenteItem.quantite).desc())\
        .limit(10).all()


        return jsonify({
            "ca_total": ca_total,
            "nb_ventes": nb_ventes,
            "panier_moyen": panier_moyen,
            "evolution": evolution,
            
            "ventes_par_categorie": [{"categorie": c.categorie, "total": c.total} for c in ventes_par_categorie],
            "ventes_par_paiement": [{"mode": p[0], "total": p[1]} for p in ventes_par_paiement]
        })

        
    except Exception as e:
        print(f"Erreur stats ventes: {e}")
        return jsonify({"error": str(e)}), 500

@statistiques_bp.route('/api/rapport/rentabilite')
@login_required
def rapport_rentabilite():
    """Analyse de rentabilité globale pour ventes multi-produits"""
    try:
        ventes = Vente.query.filter_by(statut='confirmée').all()
        
        ca_total = 0
        cout_total = 0
        produits_data = {}

        for vente in ventes:
            for item in vente.items:  # Boucler sur tous les items de la vente
                produit = item.produit  # Relation vers Produit
                if not produit:
                    continue
                
                montant_item = item.prix_unitaire * item.quantite * (1 - (item.remise or 0)/100)
                ca_total += montant_item
                cout_vente = item.quantite * produit.prix_achat
                cout_total += cout_vente

                # Collecte par produit pour le top rentable
                if produit.id not in produits_data:
                    produits_data[produit.id] = {
                        'nom': produit.nom,
                        'quantite': 0,
                        'ca': 0,
                        'cout': 0
                    }
                produits_data[produit.id]['quantite'] += item.quantite
                produits_data[produit.id]['ca'] += montant_item
                produits_data[produit.id]['cout'] += cout_vente
        
        benefice_brut = ca_total - cout_total
        marge_brute_pct = (benefice_brut / ca_total * 100) if ca_total else 0

        top_rentables = sorted(
            [
                {
                    'nom': p['nom'],
                    'quantite': p['quantite'],
                    'ca': p['ca'],
                    'cout': p['cout'],
                    'benefice': p['ca'] - p['cout'],
                    'marge': ((p['ca'] - p['cout']) / p['ca'] * 100) if p['ca'] else 0
                } for p in produits_data.values() if p['quantite'] > 0
            ],
            key=lambda x: x['benefice'],
            reverse=True
        )[:5]

        return jsonify({
            'resume': {
                'ca_total': round(ca_total, 2),
                'cout_total': round(cout_total, 2),
                'benefice_brut': round(benefice_brut, 2),
                'marge_brute': round(marge_brute_pct, 2)
            },
            'top_rentables': top_rentables
        })
    
    except Exception as e:
        print(f"Erreur dans rapport rentabilité: {e}")
        return jsonify({
            'resume': {
                'ca_total': 0,
                'cout_total': 0,
                'benefice_brut': 0,
                'marge_brute': 0
            },
            'top_rentables': []
        }), 500

@statistiques_bp.route('/api/stats/depenses')
@login_required
def stats_depenses():
    """Récupère le total des dépenses avec filtrage par période"""
    try:
        periode = request.args.get('periode', 'mois')
        now = datetime.now()

        # Calcul de la période
        if periode == 'jour':
            debut = datetime(now.year, now.month, now.day)
        elif periode == 'semaine':
            debut = now - timedelta(days=now.weekday())
        elif periode == 'mois':
            debut = datetime(now.year, now.month, 1)
        elif periode == 'annee':
            debut = datetime(now.year, 1, 1)
        else:
            debut = datetime(now.year, now.month, 1)

        # Total des dépenses pour la période
        total_depenses = db.session.query(func.sum(Depense.montant)).filter(
            Depense.date >= debut
        ).scalar() or 0

        # Nombre de dépenses
        nb_depenses = db.session.query(func.count(Depense.id)).filter(
            Depense.date >= debut
        ).scalar() or 0

        # Détail des dépenses
        depenses_list = db.session.query(
            Depense.id,
            Depense.libelle,
            Depense.montant,
            Depense.date
        ).filter(
            Depense.date >= debut
        ).order_by(Depense.date.desc()).all()

        depenses_detail = [
            {
                'id': d.id,
                'libelle': d.libelle,
                'montant': round(d.montant, 2),
                'date': d.date.strftime('%Y-%m-%d %H:%M') if d.date else ''
            }
            for d in depenses_list
        ]

        return jsonify({
            'total_depenses': round(total_depenses, 2),
            'nb_depenses': nb_depenses,
            'depenses': depenses_detail
        })

    except Exception as e:
        print(f"❌ Erreur stats dépenses: {e}")
        return jsonify({
            'total_depenses': 0,
            'nb_depenses': 0,
            'depenses': [],
            'error': str(e)
        }), 500
 
@statistiques_bp.route('/api/stats/clients')
@login_required
def stats_clients():
    """Statistiques des clients"""
    try:
        # Top clients par achats
        top_clients = db.session.query(
            Client.id,
            Client.nom,
            Client.prenom,
            func.count(Vente.id).label('nb_achats'),
            func.sum(Vente.montant_total).label('total')
        ).join(Vente, Vente.client_id == Client.id)\
         .filter(Vente.statut == 'confirmée')\
         .group_by(Client.id)\
         .order_by(func.sum(Vente.montant_total).desc())\
         .limit(10).all()

        top_clients_list = [
            {
                "id": c.id,
                "nom": f"{c.nom} {c.prenom}" if c.prenom else c.nom,
                "nb_achats": c.nb_achats or 0,
                "total": c.total or 0
            }
            for c in top_clients
        ]

        # Nouveaux clients par mois (6 derniers mois)
        nouveaux_par_mois = []
        for i in range(5, -1, -1):
            date_mois = datetime.now() - timedelta(days=30*i)
            debut_mois = datetime(date_mois.year, date_mois.month, 1)
            if date_mois.month == 12:
                fin_mois = datetime(date_mois.year + 1, 1, 1)
            else:
                fin_mois = datetime(date_mois.year, date_mois.month + 1, 1)
            
            nb_nouveaux = db.session.query(func.count(Client.id)).filter(
                Client.date_creation >= debut_mois,
                Client.date_creation < fin_mois
            ).scalar() or 0
            
            nouveaux_par_mois.append({
                "mois": debut_mois.strftime("%m/%Y"),
                "nombre": nb_nouveaux
            })

        return jsonify({
            "top_clients": top_clients_list,
            "nouveaux_par_mois": nouveaux_par_mois,
            "total_clients": len(top_clients_list)
        })

    except Exception as e:
        print(f"❌ Erreur stats clients: {e}")
        return jsonify({
            "top_clients": [],
            "nouveaux_par_mois": [],
            "total_clients": 0,
            "error": str(e)
        }), 500

@statistiques_bp.route('/api/stats/produits')
@login_required
def stats_produits():
    """
    Statistiques par produit :
    - Quantité vendue
    - Chiffre d'affaires généré
    """
    try:
        # Début du mois pour filtrer par mois courant (optionnel, à adapter)
        today = datetime.now()
        start_of_month = datetime(today.year, today.month, 1)

        # Requête : top 10 produits vendus par quantité
        top_produits = (
            db.session.query(
                Produit.nom,
                func.sum(VenteItem.quantite).label('quantite'),
                func.sum(VenteItem.montant_total).label('ca')
            )
            .join(VenteItem, Produit.id == VenteItem.produit_id)
            .join(Vente, Vente.id == VenteItem.vente_id)
            .filter(
                Vente.statut == 'confirmée',
                Vente.date_vente >= start_of_month  # ou supprime si tu veux toutes les ventes
            )
            .group_by(Produit.id)
            .order_by(func.sum(VenteItem.quantite).desc())  # Tri par quantité
            .limit(10)
            .all()
        )

        # Préparer la réponse JSON
        result = [
            {
                'nom': p[0],
                'quantite': int(p[1] or 0),
                'ca': float(p[2] or 0)
            }
            for p in top_produits
        ]

        return jsonify({'top_ventes': result})

    except Exception as e:
        print(f"Erreur stats produits: {e}")
        return jsonify({'error': str(e)}), 500


def stats_top_produits():
    """Statistiques des produits - TOP 10 produits vendus"""
    try:
        # Top produits vendus - Version corrigée
        top_ventes = db.session.query(
            Produit.id,
            Produit.nom,
            func.sum(Vente.quantite).label('quantite'),
            func.sum(Vente.montant_total).label('ca')
        ).join(Vente, Vente.produit_id == Produit.id)\
            .filter(Vente.statut == 'confirmée')\
            .group_by(Produit.id)\
            .order_by(func.sum(Vente.montant_total).desc())\
            .limit(10).all()

        top_ventes_list = [
            {
                "id": p.id,
                "nom": p.nom,
                "quantite": p.quantite or 0,
                "ca": p.ca or 0
            }
            for p in top_ventes
        ]

        return jsonify({
            "top_ventes": top_ventes_list,
            "total_produits": len(top_ventes_list)
        })

    except Exception as e:
        print(f"❌ Erreur stats produits: {e}")
        return jsonify({
            "top_ventes": [],
            "total_produits": 0,
            "error": str(e)
        }), 500
    

from flask import send_file
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

@statistiques_bp.route('/api/export/ventes/excel', methods=['GET'])
@login_required
def export_ventes_stats_excel():
    """Exporte les ventes statistiques en Excel"""
    try:
        # Filtre optionnel par période
        periode = request.args.get('periode', 'mois')
        now = datetime.now()
        if periode == 'jour':
            debut = datetime(now.year, now.month, now.day)
        elif periode == 'mois':
            debut = datetime(now.year, now.month, 1)
        elif periode == 'annee':
            debut = datetime(now.year, 1, 1)
        else:
            debut = datetime(now.year, now.month, 1)

        ventes = Vente.query.filter(
            Vente.statut == 'confirmée',
            Vente.date_vente >= debut
        ).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventes Stats"

        headers = ['ID', 'Facture', 'Date', 'Client', 'Montant', 'Mode Paiement', 'Devise']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, vente in enumerate(ventes, 2):
            ws.cell(row=row, column=1, value=vente.id)
            ws.cell(row=row, column=2, value=vente.numero_facture)
            ws.cell(row=row, column=3, value=vente.date_vente.strftime('%Y-%m-%d %H:%M') if vente.date_vente else '')
            ws.cell(row=row, column=4, value=f"{vente.client.nom} {vente.client.prenom}" if vente.client else "Client anonyme")
            ws.cell(row=row, column=5, value=float(vente.montant_total))
            ws.cell(row=row, column=6, value=vente.mode_paiement)
            ws.cell(row=row, column=7, value=vente.devise)

        # Sauvegarde dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"ventes_stats_{periode}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"❌ Erreur export ventes stats: {e}")
        return jsonify({'error': str(e)}), 500


@statistiques_bp.route('/api/export/top-produits/excel', methods=['GET'])
@login_required
def export_top_produits_excel():
    """Exporte le top 10 produits vendus en Excel"""
    try:
        top_produits = (
            db.session.query(
                Produit.nom,
                func.sum(VenteItem.quantite).label('quantite'),
                func.sum(VenteItem.montant_total).label('ca')
            )
            .join(VenteItem, Produit.id == VenteItem.produit_id)
            .join(Vente, Vente.id == VenteItem.vente_id)
            .filter(Vente.statut == 'confirmée')
            .group_by(Produit.id)
            .order_by(func.sum(VenteItem.quantite).desc())
            .limit(10)
            .all()
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Top Produits"

        headers = ['Produit', 'Quantité Vendue', 'Chiffre Affaires']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, p in enumerate(top_produits, 2):
            ws.cell(row=row, column=1, value=p[0])
            ws.cell(row=row, column=2, value=int(p[1] or 0))
            ws.cell(row=row, column=3, value=float(p[2] or 0))

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="top_produits.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"❌ Erreur export top produits: {e}")
        return jsonify({'error': str(e)}), 500
