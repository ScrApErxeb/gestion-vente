from flask import Blueprint, request, jsonify, render_template
from flask_login import login_required, current_user
from sqlalchemy import or_
from models import Vente, Produit, Client, MouvementStock, db, VenteItem, ParametreSysteme, Paiement, MouvementCaisse
from datetime import datetime
import random
from utils.export import exporter_facture_pdf

ventes_bp = Blueprint('ventes', __name__)

@ventes_bp.route('/ventes')
@login_required
def ventes_page():
    return render_template('ventes.html')

from flask import send_file

@ventes_bp.route('/api/export/facture/<int:id>', methods=['GET'])
@login_required
def export_facture(id):
    # Charger la vente avec les relations produits
    vente = Vente.query.options(
        db.joinedload(Vente.items).joinedload(VenteItem.produit),
        db.joinedload(Vente.client)
    ).get_or_404(id)

    pdf_buffer = exporter_facture_pdf(vente)

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"facture_{vente.numero_facture}.pdf",
        mimetype='application/pdf'
    )


@ventes_bp.route('/api/ventes', methods=['GET'])
@login_required
def get_ventes():
    try:
        print("🔍 Début get_ventes...")

        # Récupération des filtres
        date_debut = request.args.get('date_debut')
        date_fin = request.args.get('date_fin')
        client_id = request.args.get('client_id')
        statut = request.args.get('statut')

        print(f"📅 Filtres - date_debut: {date_debut}, date_fin: {date_fin}")

        # Chargement des ventes avec relations
        query = Vente.query.options(
            db.joinedload(Vente.items).joinedload(VenteItem.produit),
            db.joinedload(Vente.client)
        )

        # Application des filtres
        if date_debut:
            try:
                date_obj = datetime.strptime(date_debut, '%Y-%m-%d')
                query = query.filter(Vente.date_vente >= date_obj)
            except ValueError as e:
                return jsonify({'error': 'Format de date début invalide. Utilisez YYYY-MM-DD'}), 400

        if date_fin:
            try:
                date_obj = datetime.strptime(date_fin, '%Y-%m-%d')
                date_obj = date_obj.replace(hour=23, minute=59, second=59)
                query = query.filter(Vente.date_vente <= date_obj)
            except ValueError as e:
                return jsonify({'error': 'Format de date fin invalide. Utilisez YYYY-MM-DD'}), 400

        if client_id:
            query = query.filter(Vente.client_id == client_id)
        if statut:
            query = query.filter(Vente.statut == statut)

        ventes = query.order_by(Vente.date_vente.desc()).all()
        print(f"✅ {len(ventes)} ventes trouvées")

        # Formatage des ventes
        ventes_data = []
        for v in ventes:
            try:
                vente_dict = {
                    'id': v.id,
                    'numero_facture': v.numero_facture,
                    'date_vente': v.date_vente.isoformat() if v.date_vente else None,
                    'mode_paiement': v.mode_paiement,
                    'devise': v.devise,
                    'statut': v.statut,
                    'statut_paiement': v.statut_paiement,
                    'notes': v.notes,
                    'client_id': v.client_id,
                    'montant_total': float(v.montant_total) if v.montant_total else 0,
                    'client': f"{v.client.nom} {v.client.prenom}".strip() if v.client else "Client anonyme",
                    'items': [
                        {
                            'produit_id': item.produit.id,
                            'produit_nom': item.produit.nom,
                            'quantite': item.quantite,
                            'prix_unitaire': float(item.prix_unitaire),
                            'remise': float(item.remise or 0),
                            'montant_total': float(item.prix_unitaire * item.quantite) - float(item.remise or 0)
                        }
                        for item in v.items
                    ]
                }
                ventes_data.append(vente_dict)
            except Exception as e:
                print(f"❌ Erreur formatage vente {v.id}: {e}")
                # fallback basique
                ventes_data.append({
                    'id': v.id,
                    'numero_facture': v.numero_facture,
                    'client': 'Erreur chargement',
                    'montant_total': float(v.montant_total) if v.montant_total else 0
                })

        print(f"🎯 {len(ventes_data)} ventes formatées")
        return jsonify(ventes_data)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500




@ventes_bp.route('/api/ventes/<int:id>/annuler', methods=['PUT'])
@login_required
def annuler_vente(id):
    try:
        vente = Vente.query.get_or_404(id)
        
        if vente.statut == 'annulée':
            return jsonify({'error': 'Vente déjà annulée'}), 400
        
        # Restaurer le stock pour chaque produit de la vente
        for item in vente.items:
            produit = item.produit
            if produit:
                stock_avant = produit.stock_actuel
                produit.stock_actuel += item.quantite
                
                # Créer mouvement de stock d'annulation
                mouvement = MouvementStock(
                    produit_id=produit.id,
                    type_mouvement='entrée',
                    quantite=item.quantite,
                    quantite_avant=stock_avant,
                    quantite_apres=produit.stock_actuel,
                    motif=f'Annulation vente {vente.numero_facture}',
                    utilisateur=current_user.username
                )
                db.session.add(mouvement)
        
        vente.statut = 'annulée'
        db.session.commit()
        
        return jsonify({'message': 'Vente annulée avec succès'})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur annuler_vente: {e}")
        return jsonify({'error': str(e)}), 500


@ventes_bp.route('/api/ventes', methods=['POST'])
@login_required
def create_vente_api():
    try:
        data = request.get_json()
        client_id = data.get('client_id')
        devise = data.get('devise', 'XOF')
        mode_paiement = data.get('mode_paiement', 'espèces')
        notes = data.get('notes', '')
        items = data.get('items', [])

        if not items:
            return jsonify({'error': 'Aucun produit dans la vente'}), 400

        # Génération d'un numéro de facture unique
        now = datetime.now()
        suffix = random.randint(1000, 9999)
        numero_facture = f"FACT-{now.strftime('%Y%m%d')}-{suffix}"

        # Création de la vente
        vente = Vente(
            numero_facture=numero_facture,
            client_id=client_id,
            devise=devise,
            mode_paiement=mode_paiement,
            notes=notes,
            statut='confirmée',
            statut_paiement='payé'
        )

        total_vente = 0
        for item in items:
            produit_id = item['produit_id']
            quantite = item['quantite']
            prix_unitaire = item['prix_unitaire']
            remise = item.get('remise', 0)

            montant_total_item = quantite * prix_unitaire * (1 - remise/100)
            total_vente += montant_total_item

            # Création de l'item de vente
            vente_item = VenteItem(
                produit_id=produit_id,
                quantite=quantite,
                prix_unitaire=prix_unitaire,
                remise=remise,
                montant_total=montant_total_item
            )
            vente.items.append(vente_item)

            # Mise à jour du stock
            produit = Produit.query.get(produit_id)
            if produit:
                if produit.stock_actuel < quantite:
                    return jsonify({'error': f'Stock insuffisant pour {produit.nom}'}), 400

                stock_avant = produit.stock_actuel
                produit.stock_actuel -= quantite

                # Création d'un mouvement de stock
                mouvement = MouvementStock(
                    produit_id=produit.id,
                    type_mouvement='sortie',
                    quantite=quantite,
                    quantite_avant=stock_avant,
                    quantite_apres=produit.stock_actuel,
                    motif=f'Vente {numero_facture}',
                    utilisateur=current_user.username
                )
                db.session.add(mouvement)

        vente.montant_total = total_vente

        db.session.add(vente)

        # Si la vente est payée immédiatement, créer un paiement et créditer la caisse
        # (statut_paiement peut être 'payé' ou 'crédit' selon l'usage)
        if vente.statut_paiement == 'payé' or mode_paiement.lower() != 'crédit':
            # flush to get vente.id
            db.session.flush()

            # Créer paiement
            paiement = Paiement(vente_id=vente.id, montant=vente.montant_total, mode_paiement=mode_paiement)
            db.session.add(paiement)

            # Mettre à jour solde caisse
            solde_avant = ParametreSysteme.get_value("solde_caisse", 0)
            solde_apres = solde_avant + (vente.montant_total or 0)
            ParametreSysteme.set_value("solde_caisse", solde_apres, type_valeur="number")

            # Historiser mouvement de caisse
            utilisateur = getattr(current_user, 'username', None) if current_user and hasattr(current_user, 'username') else None
            mouvement = MouvementCaisse(
                type='encaisse',
                montant=vente.montant_total,
                paiement_id=paiement.id,
                vente_id=vente.id,
                utilisateur=utilisateur,
                solde_avant=solde_avant,
                solde_apres=solde_apres,
                notes='Paiement automatique lors création vente'
            )
            db.session.add(mouvement)

        db.session.commit()

        return jsonify({'message': 'Vente enregistrée', 'vente_id': vente.id, 'numero_facture': numero_facture}), 201

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur create_vente_api: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@ventes_bp.route('/api/clients', methods=['GET'])
@login_required
def get_clients():
    try:
        clients = Client.query.filter_by(actif=True).all()
        
        clients_data = []
        for c in clients:
            clients_data.append({
                'id': c.id,
                'nom': c.nom or '',
                'prenom': c.prenom or '',
                'telephone': c.telephone or '',
                'email': c.email or ''
            })
        
        print(f"✅ {len(clients_data)} clients chargés")
        return jsonify(clients_data)
    except Exception as e:
        print(f"❌ Erreur get_clients: {e}")
        return jsonify({'error': str(e)}), 500
    

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from flask import send_file

@ventes_bp.route('/api/export/ventes/excel', methods=['GET'])
@login_required
def export_ventes_excel():
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')

    if not date_debut or not date_fin:
        return jsonify({'error': 'Les paramètres date_debut et date_fin sont requis'}), 400

    try:
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except ValueError:
        return jsonify({'error': 'Format de date invalide. Utilisez YYYY-MM-DD'}), 400

    # Récupérer les ventes dans la période
    ventes = Vente.query.filter(
        Vente.date_vente >= date_debut_obj,
        Vente.date_vente <= date_fin_obj
    ).all()

    # Créer Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    headers = ['ID', 'Numéro Facture', 'Date', 'Client', 'Montant Total', 'Devise', 'Statut']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row, vente in enumerate(ventes, 2):
        ws.cell(row=row, column=1, value=vente.id)
        ws.cell(row=row, column=2, value=vente.numero_facture)
        ws.cell(row=row, column=3, value=vente.date_vente.strftime('%Y-%m-%d %H:%M') if vente.date_vente else '')
        ws.cell(row=row, column=4, value=f"{vente.client.nom} {vente.client.prenom}" if vente.client else 'Client anonyme')
        ws.cell(row=row, column=5, value=float(vente.montant_total))
        ws.cell(row=row, column=6, value=vente.devise)
        ws.cell(row=row, column=7, value=vente.statut)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'ventes_{date_debut}_a_{date_fin}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
