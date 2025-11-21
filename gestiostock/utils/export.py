"""
Fonctions d'export (PDF, Excel)
"""
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def exporter_ventes_pdf(ventes, filename=None):
    """Exporte les ventes en PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Rapport des Ventes", styles['Title']))
    elements.append(Spacer(1, 12))
    
    # Préparer les données du tableau
    data = [['Facture', 'Client', 'Produit', 'Quantité', 'Montant', 'Date']]
    
    for vente in ventes:
        data.append([
            vente.numero_facture,
            vente.client.nom if vente.client else 'Anonyme',
            vente.produit.nom,
            str(vente.quantite),
            f"{vente.montant_total:,.0f} F",
            vente.date_vente.strftime('%d/%m/%Y')
        ])
    
    # Créer le tableau
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def exporter_facture_pdf(vente):
    """Exporte une facture PDF pour une vente (simple ou multi-produits)"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=18
    )
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=18, spaceAfter=12))

    # Titre
    elements.append(Paragraph(f"Facture N° {vente.numero_facture}", styles['CenterTitle']))

    # Infos client et paiement
    client_nom = vente.client_acheteur.nom if vente.client_acheteur else "Client anonyme"
    client_info = f"""
    <b>Client:</b> {client_nom}<br/>
    <b>Date:</b> {vente.date_vente.strftime('%d/%m/%Y %H:%M')}<br/>
    <b>Mode paiement:</b> {vente.mode_paiement}<br/>
    <b>Devise:</b> {vente.devise}
    """
    elements.append(Paragraph(client_info, styles['Normal']))
    elements.append(Spacer(1, 12))

    # Tableau produits
    data = [['Produit', 'Quantité', 'Prix Unitaire', 'Remise (%)', 'Montant']]
    montant_total = 0

    # Vente multi-produits
    if hasattr(vente, 'items_vente') and vente.items_vente:
        for item in vente.items_vente:
            produit_nom = item.produit.nom if item.produit else "Produit inconnu"
            montant = (item.prix_unitaire * item.quantite) * (1 - (item.remise or 0) / 100)
            montant_total += montant
            data.append([
                produit_nom,
                str(item.quantite),
                f"{item.prix_unitaire:,.2f} F",
                f"{(item.remise or 0):.2f}%",
                f"{montant:,.2f} F"
            ])
    else:
        # Vente simple
        produit_nom = vente.produit_vendu.nom if vente.produit_vendu else "Produit inconnu"
        montant_total = vente.montant_total
        data.append([
            produit_nom,
            str(vente.quantite),
            f"{vente.prix_unitaire:,.2f} F",
            f"{vente.remise:.2f}%",
            f"{montant_total:,.2f} F"
        ])

    # Création du tableau
    table = Table(data, colWidths=[200, 60, 80, 60, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Total
    elements.append(Paragraph(f"<b>Total à payer: {montant_total:,.2f} {vente.devise}</b>", styles['Heading2']))

    # Notes éventuelles
    if getattr(vente, 'notes', None):
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"<b>Notes:</b> {vente.notes}", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer



def exporter_produits_excel(produits, filename=None):


    """Exporte les produits en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"
    
    # En-têtes
    headers = ['Référence', 'Nom', 'Catégorie', 'Prix Achat', 'Prix Vente', 'Stock', 'Stock Min', 'Fournisseur']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Données
    for row, produit in enumerate(produits, 2):
        ws.cell(row=row, column=1, value=produit.reference)
        ws.cell(row=row, column=2, value=produit.nom)
        ws.cell(row=row, column=3, value=produit.categorie.nom if produit.categorie else '')
        ws.cell(row=row, column=4, value=produit.prix_achat)
        ws.cell(row=row, column=5, value=produit.prix_vente)
        ws.cell(row=row, column=6, value=produit.stock_actuel)
        ws.cell(row=row, column=7, value=produit.stock_min)
        ws.cell(row=row, column=8, value=produit.fournisseur.nom if produit.fournisseur else '')
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer